"""Excel / CSV 题库导入：解析、逐行校验、归一化。

规则原样来自单文件 HTML 版的 normType / parseOpts / normAns / handleRows，
但不再写死在代码里 —— 题型关键词、判断题答案的各种写法、选项分隔符、
最多认几个选项，全部读 config.yaml 的 import 段，改配置即改行为。
"""

import csv
import hashlib
import io
import re

from openpyxl import Workbook, load_workbook

from ..siteconfig import site

_IMG_RE = re.compile(r"^(data:image|https?://|/uploads/)")

# 内部记号：这张表没有题干列
HEADLESS = "__headless__"


def _opt_regex() -> re.Pattern:
    """按配置里的分隔符拼出 "A.内容" 的匹配式。"""
    seps = re.escape(site.import_.option_separators)
    return re.compile(rf"^([A-Za-z])\s*[{seps}]?\s*(.+)$")


def stem_hash(stem: str) -> str:
    """题干去掉所有空白后取哈希，用来查重。"""
    return hashlib.sha256(re.sub(r"\s+", "", stem).encode("utf-8")).hexdigest()


def norm_type(value) -> str:
    """题型识别。题型栏里含有配置的关键词之一，就算这种题型。"""
    s = str(value or "").strip()
    for qtype, keywords in site.import_.type_keywords.items():
        if any(k in s for k in keywords):
            return qtype
    return ""


def parse_options(value, qtype: str) -> list[tuple[str, str]]:
    """把"可选项"一栏拆成 [(标签, 内容)]。判断题自动补两个选项。"""
    if qtype == "判断题":
        return [(o[0], o[1]) for o in site.import_.judge_options]

    s = "" if value is None else str(value).strip()
    if not s:
        return []

    pattern = _opt_regex()
    limit = site.import_.max_options
    labels = "ABCDEFGHIJ"
    out: list[tuple[str, str]] = []
    for i, part in enumerate(x.strip() for x in re.split(r"[\r\n]+", s) if x.strip()):
        m = pattern.match(part)
        if m:
            out.append((m.group(1).upper(), m.group(2).strip()))
        else:
            out.append((labels[i] if i < len(labels) else str(i + 1), part))
    return out[:limit]


def norm_answer(value, qtype: str) -> str:
    """判断题答案归一成标准写法（√/对/T/是 → 正确）。其他题型原样保留。"""
    s = "" if value is None else str(value).strip()
    if qtype == "判断题":
        for standard, variants in site.import_.judge_answers.items():
            if any(s.lower() == str(v).lower() for v in variants):
                return standard
    return s


def norm_image(value) -> str | None:
    s = str(value or "").strip()
    return s if s and _IMG_RE.match(s) else None


def _read_csv(content: bytes) -> list[list[str]]:
    """CSV 兼容 UTF-8(BOM) 和 GBK —— 老师用 Excel 另存的 CSV 多半是 GBK。"""
    for encoding in ("utf-8-sig", "gbk", "utf-8"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("这个 CSV 的编码认不出来，建议另存为 .xlsx 再传")
    return [row for row in csv.reader(io.StringIO(text))]


def _validate_rows(grid: list[list], sheet_name: str, valid_scopes: set[str]):
    """校验一张表的所有行，返回 (合格题目, 错误明细, 扫描到的数据行数)。"""
    good: list[dict] = []
    errors: list[dict] = []
    rows_seen = 0

    if not grid:
        return good, errors, rows_seen

    headers = site.import_.headers
    head = [str(h or "").strip() for h in grid[0]]
    idx = {h: (head.index(h) if h in head else -1) for h in headers}

    if idx.get("题型", -1) < 0 or idx.get("题干", -1) < 0:
        # 先打个记号。说明页、对照表这类表本来就没有题干列，
        # 只有整本工作簿都读不出题时才算错误，否则用自带模板导入会平白多一条“失败”。
        errors.append({"sheet": sheet_name, "row": 1, "reason": HEADLESS})
        return good, errors, rows_seen

    for i, row in enumerate(grid[1:], start=2):
        def get(col: str):
            j = idx.get(col, -1)
            return row[j] if 0 <= j < len(row) else ""

        stem = str(get("题干") or "").strip()
        if not stem:
            continue  # 空行忽略，不算错误
        rows_seen += 1

        qtype = norm_type(get("题型"))
        if not qtype:
            errors.append(
                {"sheet": sheet_name, "row": i,
                 "reason": f"题型「{str(get('题型') or '')}」无法识别"}
            )
            continue

        scope = str(get("知识范围") or "").strip()
        if scope not in valid_scopes:
            errors.append(
                {"sheet": sheet_name, "row": i,
                 "reason": f"知识范围「{scope or '空'}」不在允许的范围之内"}
            )
            continue

        options = parse_options(get("可选项"), qtype)
        answer = norm_answer(get("答案"), qtype)

        if qtype == "选择题":
            if len(options) < 2:
                errors.append(
                    {"sheet": sheet_name, "row": i, "reason": "选择题至少要有两个可选项"}
                )
                continue
            labels = {lb for lb, _ in options}
            if answer and answer.upper() not in labels:
                errors.append(
                    {"sheet": sheet_name, "row": i,
                     "reason": f"答案「{answer}」不在可选项 {'/'.join(sorted(labels))} 之内"}
                )
                continue
            answer = answer.upper()

        good.append(
            {
                "type": qtype,
                "stem": stem,
                "answer": answer,
                "scope": scope,
                "source": site.bank.default_source,
                "image_url": norm_image(get("图片")),
                "options": options,
                "sheet": sheet_name,
                "row": i,
            }
        )

    return good, errors, rows_seen


def parse_upload(content: bytes, filename: str, valid_scopes: set[str]):
    """解析上传的 xlsx 或 csv，返回 (合格题目, 错误明细, 数据行数)。"""
    if filename.lower().endswith(".csv"):
        good, errors, seen = _validate_rows(_read_csv(content), "CSV", valid_scopes)
        for e in errors:  # CSV 只有一张表，没表头就是真的没表头
            if e["reason"] == HEADLESS:
                e["reason"] = "找不到「题型」「题干」表头，整表跳过"
        return good, errors, seen

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    all_good, all_errors, total = [], [], 0
    headless: list[str] = []
    usable = 0

    for sheet in wb.worksheets:  # 多张工作表一起读
        grid = [list(r) for r in sheet.iter_rows(values_only=True)]
        if not grid:
            continue
        good, errors, seen = _validate_rows(grid, sheet.title, valid_scopes)
        if errors and errors[0]["reason"] == HEADLESS:
            headless.append(sheet.title)
            continue
        usable += 1
        all_good += good
        all_errors += errors
        total += seen

    if usable == 0:  # 整本都没有能用的表，这才是真的传错文件了
        for title in headless:
            all_errors.append(
                {"sheet": title, "row": 1, "reason": "找不到「题型」「题干」表头，整表跳过"}
            )

    wb.close()
    return all_good, all_errors, total


def build_template(scopes: list[str], types: list[str]) -> bytes:
    """生成空白模板：第一张表填写区，第二张表对照表。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "题目"
    ws.append(site.import_.headers)

    demo_scope = scopes[0] if scopes else ""
    ws.append(
        [
            "选择题",
            "在 Windows 中，用于切换当前活动窗口的快捷键是（ ）。",
            "A.Alt+Tab\nB.Ctrl+C\nC.Alt+F4\nD.Win+D",
            "A",
            demo_scope,
            "",
        ]
    )
    ws.append(["判断题", "计算机病毒是一种可以自我复制的程序。", "", "正确", demo_scope, ""])
    ws.append(["操作题", "把当前文档另存为 PDF 并命名为「作业.pdf」。", "", "略", demo_scope, ""])

    for col, w in zip("ABCDEF", [10, 60, 40, 12, 22, 30]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("对照表")
    ws2.append(["知识范围（只能填这些）", "题型"])
    for i in range(max(len(scopes), len(types))):
        ws2.append([scopes[i] if i < len(scopes) else "", types[i] if i < len(types) else ""])
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
