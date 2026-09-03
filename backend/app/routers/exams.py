"""教师端：发布考试、查成绩、导出。全部需要登录。"""

import io
import json
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Response
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..database import get_db
from ..deps import get_current_user
from ..models import Exam, ExamSubmission, Paper, User
from ..schemas import ExamCreate, ExamOut, ExamUpdate, SubmissionOut
from ..services.exam import group_items, load_items, new_token, question_stats

router = APIRouter(prefix="/api/exams", tags=["考试"])

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _to_out(exam: Exam) -> ExamOut:
    item = ExamOut.model_validate(exam)
    subs = exam.submissions
    item.submission_count = len(subs)
    item.avg_score = round(sum(s.score for s in subs) / len(subs), 1) if subs else None
    return item


def _get(db: Session, exam_id: int) -> Exam:
    exam = db.get(Exam, exam_id)
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")
    return exam


@router.post("", response_model=ExamOut, summary="把一份存档试卷发布成考试")
def create_exam(
    body: ExamCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)
):
    paper = db.get(Paper, body.paper_id)
    if not paper:
        raise HTTPException(status_code=404, detail="试卷不存在，请先在组卷页勾选「把这套卷子存档」")
    if not paper.items:
        raise HTTPException(status_code=400, detail="这份试卷没有题目")

    exam = Exam(
        paper_id=paper.id,
        title=body.title or paper.title,
        token=new_token(),
        is_open=body.is_open,
        allow_retake=body.allow_retake,
        show_score=body.show_score,
        show_answer=body.show_answer,
        created_by=user.id,
    )
    db.add(exam)
    db.commit()
    db.refresh(exam)
    return _to_out(exam)


@router.get("", response_model=list[ExamOut], summary="考试列表")
def list_exams(_: User = Depends(get_current_user), db: Session = Depends(get_db)):
    rows = db.scalars(select(Exam).order_by(Exam.id.desc())).all()
    return [_to_out(e) for e in rows]


@router.get("/{exam_id}", response_model=ExamOut, summary="考试详情")
def get_exam(exam_id: int, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return _to_out(_get(db, exam_id))


@router.patch("/{exam_id}", response_model=ExamOut, summary="改考试设置（开关、是否给学生看分数等）")
def update_exam(
    exam_id: int,
    body: ExamUpdate,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    exam = _get(db, exam_id)
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(exam, k, v)
    db.commit()
    db.refresh(exam)
    return _to_out(exam)


@router.delete("/{exam_id}", summary="删除考试（连同答卷）")
def delete_exam(exam_id: int, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    db.delete(_get(db, exam_id))
    db.commit()
    return {"ok": True}


@router.get("/{exam_id}/submissions", response_model=list[SubmissionOut], summary="成绩列表")
def list_submissions(
    exam_id: int, _: User = Depends(get_current_user), db: Session = Depends(get_db)
):
    exam = _get(db, exam_id)
    return sorted(exam.submissions, key=lambda s: (-s.score, s.student_no))


@router.get("/{exam_id}/submissions/{sub_id}", summary="某个学生的答卷明细")
def submission_detail(
    exam_id: int,
    sub_id: int,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    sub = db.get(ExamSubmission, sub_id)
    if not sub or sub.exam_id != exam_id:
        raise HTTPException(status_code=404, detail="答卷不存在")

    exam = _get(db, exam_id)
    items = {it["id"]: it for it in load_items(db, exam.paper)}
    detail = json.loads(sub.detail_json or "[]")
    for d in detail:
        it = items.get(d["id"])
        if it:
            d["stem"] = it["stem"]
            d["options"] = it["options"]
            d["scope"] = it["scope"]

    return {
        "id": sub.id,
        "student_name": sub.student_name,
        "student_class": sub.student_class,
        "student_no": sub.student_no,
        "score": sub.score,
        "right_count": sub.right_count,
        "objective_count": sub.objective_count,
        "submitted_at": sub.submitted_at,
        "detail": detail,
    }


@router.get("/{exam_id}/stats", summary="题目分析：每题正确率")
def exam_stats(exam_id: int, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    exam = _get(db, exam_id)
    items = load_items(db, exam.paper)
    subs = exam.submissions
    scores = [s.score for s in subs]
    return {
        "submission_count": len(subs),
        "avg_score": round(sum(scores) / len(scores), 1) if scores else None,
        "max_score": max(scores) if scores else None,
        "min_score": min(scores) if scores else None,
        "questions": question_stats(exam, items),
    }


@router.get("/{exam_id}/export.xlsx", summary="成绩汇总导出 Excel")
def export_scores(
    exam_id: int, _: User = Depends(get_current_user), db: Session = Depends(get_db)
):
    exam = _get(db, exam_id)
    items = load_items(db, exam.paper)
    groups = group_items(items)
    ordered = [it for g in groups for it in g["items"]]
    no_of = {it["id"]: i + 1 for i, it in enumerate(ordered)}

    wb = Workbook()

    ws = wb.active
    ws.title = "成绩汇总"
    ws.append(["姓名", "班级", "学号", "得分(百分制)", "答对", "客观题数", "交卷时间"])
    for s in sorted(exam.submissions, key=lambda x: (-x.score, x.student_no)):
        ws.append(
            [
                s.student_name,
                s.student_class,
                s.student_no,
                s.score,
                s.right_count,
                s.objective_count,
                s.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if s.submitted_at else "",
            ]
        )
    for col, w in zip("ABCDEFG", [12, 14, 14, 14, 8, 10, 20]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    # 第二张表：每题正确率，老师用来讲评
    ws2 = wb.create_sheet("题目分析")
    ws2.append(["题号", "题型", "知识范围", "题干", "答案", "答对", "答错", "未答", "正确率%"])
    for st in question_stats(exam, items):
        ws2.append(
            [
                no_of.get(st["id"], ""),
                st["type"],
                st["scope"],
                st["stem"],
                st["answer"],
                st["right"],
                st["wrong"],
                st["blank"],
                st["accuracy"] if st["scorable"] else "",
            ]
        )
    for col, w in zip("ABCDEFGHI", [6, 8, 16, 50, 10, 8, 8, 8, 10]):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    # 第三张表：操作题的原始作答，需要人工评阅
    ws3 = wb.create_sheet("操作题作答")
    ws3.append(["姓名", "班级", "学号", "题号", "学生作答", "答案要点"])
    for s in exam.submissions:
        for d in json.loads(s.detail_json or "[]"):
            if d.get("type") == "操作题":
                ws3.append(
                    [
                        s.student_name,
                        s.student_class,
                        s.student_no,
                        no_of.get(d["id"], ""),
                        d.get("mine", ""),
                        d.get("answer", ""),
                    ]
                )
    for col, w in zip("ABCDEF", [12, 14, 14, 6, 60, 60]):
        ws3.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    name = quote(f"{exam.title}_成绩.xlsx")
    return Response(
        content=buf.getvalue(),
        media_type=XLSX_MIME,
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{name}"},
    )
