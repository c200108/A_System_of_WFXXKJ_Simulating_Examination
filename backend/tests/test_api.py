"""接口层端到端测试：登录 → 建题 → 下载模板 → 上传导入 → 组卷。

跑的是真实的 FastAPI 应用和真实的 SQLite 库（临时目录里的），不是 mock。
"""

XLSX_MAGIC = b"PK\x03\x04"


def test_health(client):
    assert client.get("/api/health").json() == {"status": "ok"}


def test_login_rejects_wrong_password(client):
    res = client.post("/api/auth/login", data={"username": "admin", "password": "wrong"})
    assert res.status_code == 400


def test_endpoints_require_login(client):
    assert client.get("/api/questions").status_code == 401


def test_me(client, auth):
    body = client.get("/api/auth/me", headers=auth).json()
    assert body["username"] == "admin"
    assert body["role"] == "admin"


def test_dicts_seeded(client, auth):
    dicts = client.get("/api/dicts", headers=auth).json()
    scopes = [d["name"] for d in dicts if d["category"] == "scope"]
    assert len(scopes) == 10
    assert "Windows系统操作" in scopes
    assert "物联网" in scopes


# ---------- 题目增删改查 ----------
def test_create_question(client, auth):
    payload = {
        "type": "选择题",
        "stem": "测试题：下列哪个是操作系统？",
        "answer": "A",
        "scope": "计算机软件",
        "options": [
            {"label": "A", "content": "Windows"},
            {"label": "B", "content": "Word"},
        ],
    }
    res = client.post("/api/questions", json=payload, headers=auth)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["code"].startswith("C")
    assert len(body["options"]) == 2


def test_duplicate_stem_rejected(client, auth):
    payload = {
        "type": "选择题",
        "stem": "测试题：下列哪个是操作系统？",  # 与上一题同题干
        "answer": "A",
        "scope": "计算机软件",
        "options": [{"label": "A", "content": "Windows"}],
    }
    assert client.post("/api/questions", json=payload, headers=auth).status_code == 409


def test_bad_scope_rejected(client, auth):
    payload = {
        "type": "选择题",
        "stem": "这道题的知识范围是瞎写的",
        "answer": "A",
        "scope": "量子力学",
        "options": [{"label": "A", "content": "甲"}],
    }
    res = client.post("/api/questions", json=payload, headers=auth)
    assert res.status_code == 400
    assert "不在十类之内" in res.json()["detail"]


def test_soft_delete_and_restore(client, auth):
    created = client.post(
        "/api/questions",
        json={"type": "判断题", "stem": "这道题一会儿要被删掉。", "answer": "正确", "scope": "人工智能"},
        headers=auth,
    ).json()
    qid = created["id"]

    assert client.delete(f"/api/questions/{qid}", headers=auth).status_code == 200
    assert client.get(f"/api/questions/{qid}", headers=auth).status_code == 404

    assert client.post(f"/api/questions/{qid}/restore", headers=auth).status_code == 200
    assert client.get(f"/api/questions/{qid}", headers=auth).status_code == 200


# ---------- 模板下载 → 上传导入 ----------
def test_template_download(client, auth):
    res = client.get("/api/imports/template", headers=auth)
    assert res.status_code == 200
    assert res.content.startswith(XLSX_MAGIC)  # 是个真正的 xlsx
    assert len(res.content) > 3000


def test_import_roundtrip(client, auth):
    """下载空白模板 → 原样传回去 → 模板里的三道示例题应该入库。"""
    template = client.get("/api/imports/template", headers=auth).content
    files = {
        "file": (
            "模板.xlsx",
            template,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    res = client.post("/api/imports/questions", files=files, headers=auth)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["success"] == 3
    assert body["failed"] == 0
    assert body["by_type"] == {"选择题": 1, "判断题": 1, "操作题": 1}

    # 同一个文件再传一次，应该全部按重复跳过，不会产生重复题
    again = client.post("/api/imports/questions", files=files, headers=auth).json()
    assert again["success"] == 0
    assert again["skipped"] == 3


def test_import_rejects_non_xlsx(client, auth):
    files = {"file": ("题库.txt", b"not an excel file", "text/plain")}
    assert client.post("/api/imports/questions", files=files, headers=auth).status_code == 400


def test_import_log_recorded(client, auth):
    logs = client.get("/api/imports/logs", headers=auth).json()
    assert len(logs) >= 2
    assert logs[0]["operator"] == "系统管理员"


# ---------- 组卷 ----------
def test_generate_paper(client, auth):
    res = client.post(
        "/api/papers/generate",
        json={"title": "单元测验", "counts": {"选择题": 2, "判断题": 1}, "save": True},
        headers=auth,
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["total"] == len(body["questions"]) == 3
    assert body["paper_id"] is not None
    assert sum(body["tally"].values()) == 3
    types = [q["type"] for q in body["questions"]]
    assert types.count("选择题") == 2 and types.count("判断题") == 1


def test_saved_paper_can_be_reopened(client, auth):
    pid = client.post(
        "/api/papers/generate",
        json={"title": "存档卷", "counts": {"选择题": 2}, "save": True},
        headers=auth,
    ).json()["paper_id"]

    reopened = client.get(f"/api/papers/{pid}", headers=auth).json()
    assert reopened["title"] == "存档卷"
    assert reopened["total"] == 2


def test_generate_with_impossible_request(client, auth):
    """要的题比库里多时，收窄到上限并给出提示，而不是报错。"""
    body = client.post(
        "/api/papers/generate", json={"counts": {"操作题": 9999}}, headers=auth
    ).json()
    assert body["total"] < 9999
    assert body["warnings"] and "上限" in body["warnings"][0]


def test_generate_nothing_is_an_error(client, auth):
    res = client.post("/api/papers/generate", json={"counts": {"选择题": 0}}, headers=auth)
    assert res.status_code == 400


# ---------- 试卷抬头、分组、打乱 ----------
def test_paper_header_and_groups(client, auth):
    body = client.post(
        "/api/papers/generate",
        json={
            "title": "2026年信息技术模拟测试（A卷）",
            "school": "昌邑市实验中学",
            "duration": "40",
            "counts": {"选择题": 2, "判断题": 2},
        },
        headers=auth,
    ).json()

    assert body["school"] == "昌邑市实验中学"
    assert body["duration"] == "40"
    assert body["code"].startswith("NO.") and len(body["code"]) == 8
    # 按大题分组，顺序跟着题型字典走
    assert [g["type"] for g in body["groups"]] == ["选择题", "判断题"]
    assert [len(g["items"]) for g in body["groups"]] == [2, 2]
    # 拉平的 questions 和分组里的题一一对应
    assert len(body["questions"]) == 4


def test_shuffled_paper_answer_still_matches_its_option(client, auth):
    body = client.post(
        "/api/papers/generate",
        json={"counts": {"选择题": 12}, "shuffle_options": True, "seed": "shuffle-check"},
        headers=auth,
    ).json()
    for q in body["questions"]:
        if q["answer"] and len(q["answer"]) == 1:
            labels = [o["label"] for o in q["options"]]
            assert q["answer"] in labels, q


def test_saved_paper_keeps_the_shuffled_order(client, auth):
    gen = client.post(
        "/api/papers/generate",
        json={"counts": {"选择题": 5}, "shuffle_options": True, "save": True},
        headers=auth,
    ).json()
    again = client.get(f"/api/papers/{gen['paper_id']}", headers=auth).json()

    before = [(q["id"], q["answer"], [o["content"] for o in q["options"]]) for q in gen["questions"]]
    after = [(q["id"], q["answer"], [o["content"] for o in q["options"]]) for q in again["questions"]]
    assert sorted(before) == sorted(after)  # 重新打开和当初印出去的一模一样


def test_no_answer_questions_are_skipped_by_default(client, auth):
    """默认跳过原卷未给答案的题：否则学生「没作答」会被判成答对。"""
    client.post(
        "/api/questions",
        json={"type": "判断题", "stem": "这道题故意不给答案，用来验证默认跳过。", "answer": "",
              "scope": "人工智能"},
        headers=auth,
    )
    body = client.post(
        "/api/papers/generate", json={"counts": {"判断题": 50}}, headers=auth
    ).json()
    assert all((q["answer"] or "").strip() for q in body["questions"])

    # 明确关掉这个开关时，没答案的题才会被抽出来
    loose = client.post(
        "/api/papers/generate",
        json={"counts": {"判断题": 50}, "require_answer": False},
        headers=auth,
    ).json()
    assert loose["total"] > body["total"]


def test_student_html_does_not_grade_answerless_questions(client, auth):
    paper = client.post(
        "/api/papers/generate",
        json={"counts": {"判断题": 50}, "require_answer": False},
        headers=auth,
    ).json()
    html = client.post("/api/papers/export/student-html", json=paper, headers=auth).content.decode()
    assert "不计分" in html


# ---------- 导出 ----------
def test_export_paper_xlsx(client, auth):
    paper = client.post(
        "/api/papers/generate", json={"counts": {"选择题": 3, "判断题": 2}}, headers=auth
    ).json()
    res = client.post("/api/papers/export/xlsx", json=paper, headers=auth)
    assert res.status_code == 200
    assert res.content.startswith(XLSX_MAGIC)
    assert "attachment" in res.headers["content-disposition"]


def test_export_student_html(client, auth):
    paper = client.post(
        "/api/papers/generate",
        json={"title": "自测卷", "school": "昌邑市实验中学", "counts": {"选择题": 3, "操作题": 1}},
        headers=auth,
    ).json()
    res = client.post("/api/papers/export/student-html", json=paper, headers=auth)
    assert res.status_code == 200

    html = res.content.decode("utf-8")
    assert html.startswith("<!doctype html>")
    assert "自测卷" in html and "昌邑市实验中学" in html
    assert "交卷判分" in html and "导出成绩单" in html
    # 题目连答案一起内嵌，学生打开就能自测（也意味着答案对学生可见）
    assert '"items"' in html and "letterOf" in html
    for q in paper["questions"]:
        assert q["stem"][:12] in html


def _read_xlsx(content: bytes) -> list[list]:
    import io

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    return [list(r) for r in wb.active.iter_rows(values_only=True)]


def test_export_bank_xlsx(client, auth):
    marker = "这道题来源是原卷，用来验证导出筛选。"
    client.post(
        "/api/questions",
        json={"type": "判断题", "stem": marker, "answer": "正确", "scope": "物联网", "source": "原卷"},
        headers=auth,
    )

    res = client.get("/api/questions/export.xlsx", headers=auth)
    assert res.status_code == 200
    assert res.content.startswith(XLSX_MAGIC)

    rows = _read_xlsx(res.content)
    assert rows[0] == ["题型", "题干", "可选项", "答案", "知识范围", "来源", "编号"]
    all_stems = [r[1] for r in rows[1:]]
    assert marker in all_stems

    # 带筛选：只导出老师自己补充的（来源=自定义），原卷那道不该出现
    custom = client.get("/api/questions/export.xlsx", params={"source": "自定义"}, headers=auth)
    custom_stems = [r[1] for r in _read_xlsx(custom.content)[1:]]
    assert marker not in custom_stems
    assert 0 < len(custom_stems) < len(all_stems)


# ---------- 来源筛选 ----------
def test_source_filter_and_stats(client, auth):
    stats = client.get("/api/questions/stats", headers=auth).json()
    assert "自定义" in stats["sources"]

    listed = client.get("/api/questions", params={"source": "自定义"}, headers=auth).json()
    assert listed["total"] > 0
    assert all(q["source"] == "自定义" for q in listed["items"])


# ---------- CSV 导入 ----------
def test_import_csv(client, auth):
    csv_text = (
        "题型,题干,可选项,答案,知识范围,图片\n"
        "判断题,CSV 导入这条题目是测试用的。,,正确,物联网,\n"
    )
    files = {"file": ("补充题.csv", csv_text.encode("utf-8"), "text/csv")}
    body = client.post("/api/imports/questions", files=files, headers=auth).json()
    assert body["success"] == 1
    assert body["by_type"] == {"判断题": 1}
