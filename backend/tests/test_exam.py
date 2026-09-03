"""正式考试：发布 → 学生取卷 → 交卷 → 后端判分 → 成绩汇总。

最要紧的一条：**答案不能出现在给学生的任何响应里**，见 test_answers_never_reach_the_student。
"""

import pytest

XLSX_MAGIC = b"PK\x03\x04"


@pytest.fixture(scope="module")
def exam(client, auth):
    """发布一场考试。题量按题库实际存量走，断言里用 total/objective 而不是写死数字。"""
    paper = client.post(
        "/api/papers/generate",
        json={
            "title": "期末测验",
            "school": "昌邑市实验中学",
            "duration": "40",
            "counts": {"选择题": 3, "判断题": 2, "操作题": 1},
            "save": True,
        },
        headers=auth,
    ).json()
    assert paper["paper_id"], "组卷要先存档才能发布考试"

    objective = [q for q in paper["questions"] if q["type"] != "操作题"]
    operation = [q for q in paper["questions"] if q["type"] == "操作题"]
    assert objective and operation, "这场考试要同时有客观题和操作题才测得出两条判分路径"

    res = client.post(
        "/api/exams",
        json={"paper_id": paper["paper_id"], "show_score": True, "show_answer": False},
        headers=auth,
    )
    assert res.status_code == 200, res.text
    return {
        "exam": res.json(),
        "paper": paper,
        "total": paper["total"],
        "objective": len(objective),
    }


# ---------- 发布 ----------
def test_publish_exam(exam):
    e = exam["exam"]
    assert e["title"] == "期末测验"
    assert len(e["token"]) >= 10  # 口令要够长，防止被猜
    assert e["is_open"] is True
    assert e["submission_count"] == 0


def test_cannot_publish_unsaved_paper(client, auth):
    res = client.post("/api/exams", json={"paper_id": 999999}, headers=auth)
    assert res.status_code == 404


def test_exam_management_needs_login(client, exam):
    assert client.get("/api/exams").status_code == 401
    assert client.get(f"/api/exams/{exam['exam']['id']}/submissions").status_code == 401


# ---------- 学生取卷 ----------
def test_student_can_open_without_login(client, exam):
    """学生不需要账号，直接凭链接进。"""
    res = client.get(f"/api/take/{exam['exam']['token']}")
    assert res.status_code == 200
    body = res.json()
    assert body["title"] == "期末测验"
    assert body["school"] == "昌邑市实验中学"
    assert body["total"] == exam["total"]
    assert [g["type"] for g in body["groups"]] == ["选择题", "判断题", "操作题"]


def test_answers_never_reach_the_student(client, exam):
    """核心安全断言：响应里既没有 answer 字段，也不含任何一道题的答案文本。"""
    raw = client.get(f"/api/take/{exam['exam']['token']}").text
    assert '"answer"' not in raw

    body = client.get(f"/api/take/{exam['exam']['token']}").json()
    for g in body["groups"]:
        for q in g["items"]:
            assert "answer" not in q
            assert "source" not in q
            assert set(q.keys()) <= {"id", "code", "type", "stem", "scope", "image_url", "options"}


def test_bad_token_is_404(client):
    assert client.get("/api/take/definitely-not-a-real-token").status_code == 404


def test_closed_exam_rejects_students(client, auth, exam):
    eid = exam["exam"]["id"]
    token = exam["exam"]["token"]
    client.patch(f"/api/exams/{eid}", json={"is_open": False}, headers=auth)
    assert client.get(f"/api/take/{token}").status_code == 403
    assert client.post(f"/api/take/{token}/submit", json={"student_name": "甲"}).status_code == 403
    client.patch(f"/api/exams/{eid}", json={"is_open": True}, headers=auth)
    assert client.get(f"/api/take/{token}").status_code == 200


# ---------- 交卷判分 ----------
def _answer_key(client, auth, exam):
    """老师视角取回带答案的卷子，测试里用来模拟「全对」的学生。"""
    paper = client.get(f"/api/papers/{exam['paper']['paper_id']}", headers=auth).json()
    return {str(q["id"]): q for q in paper["questions"]}


def test_full_marks(client, auth, exam):
    key = _answer_key(client, auth, exam)
    answers = {}
    for qid, q in key.items():
        if q["type"] == "判断题":
            answers[qid] = "A" if q["answer"] == "正确" else "B"
        elif q["type"] == "选择题":
            answers[qid] = q["answer"]
        else:
            answers[qid] = "我的操作步骤：先打开文件，再另存为。"

    res = client.post(
        f"/api/take/{exam['exam']['token']}/submit",
        json={"student_name": "张三", "student_class": "八年级一班", "student_no": "20260101",
              "answers": answers},
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["score"] == 100
    assert body["right_count"] == body["objective_count"] == exam["objective"]  # 操作题不计分
    assert body["detail"] == []  # show_answer 关着，不回传对错明细


def test_zero_marks_and_blank(client, exam):
    """全部不作答：不能因为「空 == 空」被判成答对。"""
    res = client.post(
        f"/api/take/{exam['exam']['token']}/submit",
        json={"student_name": "李四", "student_class": "八年级一班", "student_no": "20260102",
              "answers": {}},
    )
    body = res.json()
    assert body["score"] == 0
    assert body["right_count"] == 0
    assert body["objective_count"] == exam["objective"]


def test_name_is_required(client, exam):
    res = client.post(f"/api/take/{exam['exam']['token']}/submit", json={"answers": {}})
    assert res.status_code == 400


def test_same_student_cannot_submit_twice(client, exam):
    res = client.post(
        f"/api/take/{exam['exam']['token']}/submit",
        json={"student_name": "张三", "student_no": "20260101", "answers": {}},
    )
    assert res.status_code == 409
    assert "交过卷" in res.json()["detail"]


def test_retake_when_allowed(client, auth, exam):
    eid = exam["exam"]["id"]
    client.patch(f"/api/exams/{eid}", json={"allow_retake": True}, headers=auth)
    res = client.post(
        f"/api/take/{exam['exam']['token']}/submit",
        json={"student_name": "张三", "student_no": "20260101", "answers": {}},
    )
    assert res.status_code == 200
    client.patch(f"/api/exams/{eid}", json={"allow_retake": False}, headers=auth)


def test_show_score_off_hides_the_number(client, auth, exam):
    eid = exam["exam"]["id"]
    client.patch(f"/api/exams/{eid}", json={"show_score": False}, headers=auth)
    body = client.post(
        f"/api/take/{exam['exam']['token']}/submit",
        json={"student_name": "王五", "student_no": "20260103", "answers": {}},
    ).json()
    assert body["submitted"] is True
    assert body["score"] is None
    assert "老师统一公布" in body["message"]
    client.patch(f"/api/exams/{eid}", json={"show_score": True}, headers=auth)


def test_show_answer_on_returns_detail(client, auth, exam):
    eid = exam["exam"]["id"]
    client.patch(f"/api/exams/{eid}", json={"show_answer": True}, headers=auth)
    body = client.post(
        f"/api/take/{exam['exam']['token']}/submit",
        json={"student_name": "赵六", "student_no": "20260104", "answers": {}},
    ).json()
    assert len(body["detail"]) == exam["total"]
    assert any(d.get("scored") for d in body["detail"])
    client.patch(f"/api/exams/{eid}", json={"show_answer": False}, headers=auth)


# ---------- 成绩汇总 ----------
def test_teacher_sees_scores(client, auth, exam):
    subs = client.get(f"/api/exams/{exam['exam']['id']}/submissions", headers=auth).json()
    assert len(subs) >= 4
    assert subs[0]["score"] >= subs[-1]["score"]  # 按分数从高到低
    top = subs[0]
    assert top["student_name"] == "张三" and top["score"] == 100


def test_submission_detail_carries_the_stem(client, auth, exam):
    subs = client.get(f"/api/exams/{exam['exam']['id']}/submissions", headers=auth).json()
    d = client.get(
        f"/api/exams/{exam['exam']['id']}/submissions/{subs[0]['id']}", headers=auth
    ).json()
    assert d["student_name"] == "张三"
    assert all("stem" in x for x in d["detail"])
    op = [x for x in d["detail"] if x["type"] == "操作题"]
    assert op and "我的操作步骤" in op[0]["mine"]  # 操作题原文留着给老师看


def test_question_stats(client, auth, exam):
    st = client.get(f"/api/exams/{exam['exam']['id']}/stats", headers=auth).json()
    assert st["submission_count"] >= 4
    assert st["max_score"] == 100
    assert len(st["questions"]) == exam["total"]
    graded = [q for q in st["questions"] if q["scorable"]]
    assert len(graded) == exam["objective"]
    assert all(0 <= q["accuracy"] <= 100 for q in graded)


def test_export_scores_xlsx(client, auth, exam):
    res = client.get(f"/api/exams/{exam['exam']['id']}/export.xlsx", headers=auth)
    assert res.status_code == 200
    assert res.content.startswith(XLSX_MAGIC)

    import io

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(res.content), read_only=True)
    assert wb.sheetnames == ["成绩汇总", "题目分析", "操作题作答"]
    rows = [list(r) for r in wb["成绩汇总"].iter_rows(values_only=True)]
    assert rows[0] == ["姓名", "班级", "学号", "得分(百分制)", "答对", "客观题数", "交卷时间"]
    assert any(r[0] == "张三" and r[3] == 100 for r in rows[1:])


def test_delete_exam_removes_submissions(client, auth):
    paper = client.post(
        "/api/papers/generate", json={"counts": {"选择题": 2}, "save": True}, headers=auth
    ).json()
    e = client.post("/api/exams", json={"paper_id": paper["paper_id"]}, headers=auth).json()
    client.post(
        f"/api/take/{e['token']}/submit", json={"student_name": "临时", "answers": {}}
    )
    assert client.delete(f"/api/exams/{e['id']}", headers=auth).status_code == 200
    assert client.get(f"/api/take/{e['token']}").status_code == 404
